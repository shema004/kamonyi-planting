# data_loader.py
import numpy as np
import pandas as pd
from openpyxl import load_workbook
from pathlib import Path

RAINFALL_SECTORS = [
    "Gacurabwenge","Karama","Kayenzi","Kayumbu","Mugina",
    "Musambira","Nyamiyaga","Nyarubaka","Rugarika","Rukoma","Runda"
]
TEMP_SECTORS = RAINFALL_SECTORS + ["Ngamba"]

def _safe_float(val):
    if val is None: return np.nan
    if isinstance(val, str) and (val.startswith("#") or not val.strip()): return np.nan
    try: return float(val)
    except: return np.nan

def load_rainfall(wb):
    ws   = wb["Rainfall seasonal summary"]
    rows = list(ws.iter_rows(values_only=True))
    sector_row, season_row, field_row = rows[1], rows[2], rows[3]
    col_map, current_sector, current_season = {}, None, None
    for i, cell in enumerate(sector_row):
        if cell and isinstance(cell, str) and cell in RAINFALL_SECTORS:
            current_sector = cell
        s = season_row[i]
        if s in ("A","B"): current_season = s
        f = field_row[i]
        if f and isinstance(f, str) and current_sector:
            col_map[i] = (current_sector, current_season, f)
    records = []
    for row in rows[4:]:
        year = row[0]
        if not isinstance(year,(int,float)): continue
        year = int(year)
        if year < 2000: continue          # ignore 1999
        row_data = {}
        for ci,(sec,sea,fld) in col_map.items():
            row_data.setdefault((sec,sea),{})[fld] = _safe_float(row[ci])
        for (sec,sea),fields in row_data.items():
            records.append({"year":year,"sector":sec,"season":sea,
                "onset_day":     fields.get("Onset_DAY/YEAR", np.nan),
                "length_dekads": fields.get("Sum of Length_Dekads", np.nan),
                "total_rainfall":fields.get("Sum of Total_Rainfall", np.nan)})
    df = pd.DataFrame(records)
    df = df[df["year"]<=2025].dropna(subset=["onset_day"])
    df["onset_day"] = df["onset_day"].astype(int)
    return df.sort_values(["sector","season","year"]).reset_index(drop=True)

def _load_temp_sheet(wb, sheet_name, col_name):
    ws   = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    sector_row, season_row = rows[1], rows[2]
    col_map, current_sector = {}, None
    for i,cell in enumerate(sector_row):
        if cell and isinstance(cell,str) and cell in TEMP_SECTORS:
            current_sector = cell
        s = season_row[i]
        if s in ("A","B") and current_sector:
            col_map[i] = (current_sector, s)
    records = []
    for row in rows[3:]:
        year = row[0]
        if not isinstance(year,(int,float)): continue
        year = int(year)
        if year < 2000: continue          # ignore 1999
        for ci,(sec,sea) in col_map.items():
            records.append({"year":year,"sector":sec,"season":sea,
                            col_name:_safe_float(row[ci])})
    df = pd.DataFrame(records)
    return df[df["year"]<=2024].sort_values(["sector","season","year"]).reset_index(drop=True)


# ── Ngamba: rainfall proxy from 4 nearest sectors ─────────────────────────
# Ngamba has real temperature data but no rainfall measurements.
# Rainfall is estimated from the 4 nearest sectors by distance.
NGAMBA_PROXY_SECTORS = ["Gacurabwenge", "Karama", "Kayenzi", "Rukoma"]
NGAMBA_NOTE = (
    "Rainfall data estimated from average of nearest sectors "
    "(Gacurabwenge, Karama, Kayenzi, Rukoma). "
    "Temperature data is real (2000–2024). "
    "Recorder is collecting live data from today onwards."
)


def _add_ngamba_proxy(rf_df: pd.DataFrame,
                      mx_df: pd.DataFrame,
                      mn_df: pd.DataFrame) -> pd.DataFrame:
    """
    Build synthetic Ngamba rainfall rows by averaging 4 nearest sectors,
    then merge with Ngamba's real temperature data.
    Returns a merged DataFrame for Ngamba only.
    """
    proxy = rf_df[rf_df["sector"].isin(NGAMBA_PROXY_SECTORS)]
    ngamba_rf = (
        proxy.groupby(["year", "season"])
        .agg(
            onset_day      =("onset_day",      "mean"),
            length_dekads  =("length_dekads",   "mean"),
            total_rainfall =("total_rainfall",  "mean"),
        )
        .reset_index()
    )
    ngamba_rf["onset_day"]     = ngamba_rf["onset_day"].round(0).astype(int)
    ngamba_rf["sector"]        = "Ngamba"
    ngamba_rf["data_note"]     = NGAMBA_NOTE

    # Merge with Ngamba's real temperatures
    ngamba_mx = mx_df[mx_df["sector"] == "Ngamba"][["year","season","mean_max_temp"]]
    ngamba_mn = mn_df[mn_df["sector"] == "Ngamba"][["year","season","mean_min_temp"]]

    merged = (
        ngamba_rf
        .merge(ngamba_mx, on=["year","season"], how="left")
        .merge(ngamba_mn, on=["year","season"], how="left")
    )
    return merged

def load_all_data(path=None):
    candidates = [
        path,
        Path(__file__).parent.parent / "data" / "Overall Summary for kamonyi.xlsx",
        Path(__file__).parent.parent / "data" / "Overall_Summary_for_kamonyi.xlsx",
    ]
    xlsx = next((Path(p) for p in candidates if p and Path(p).exists()), None)
    if not xlsx:
        tried = "\n  ".join(str(c) for c in candidates if c)
        raise FileNotFoundError(
            f"Cannot find the Excel file. Tried:\n  {tried}\n\n"
            "Make sure the file is inside the data/ folder."
        )
    print(f"[data_loader] Loading: {xlsx}")
    wb = load_workbook(str(xlsx), read_only=True, data_only=True)
    rf = load_rainfall(wb)
    mx = _load_temp_sheet(wb,"Maximum temp seasonal summary","mean_max_temp")
    mn = _load_temp_sheet(wb,"Minimum temp seasonal summary","mean_min_temp")
    # Build base merged (11 sectors)
    merged = (rf.merge(mx,on=["year","sector","season"],how="left")
                .merge(mn,on=["year","sector","season"],how="left"))

    # Add Ngamba synthetic rows
    ngamba_merged = _add_ngamba_proxy(rf, mx, mn)
    # Add missing columns present in merged but not ngamba_merged
    for col in merged.columns:
        if col not in ngamba_merged.columns:
            ngamba_merged[col] = np.nan
    merged = pd.concat([merged, ngamba_merged], ignore_index=True)
    merged = merged.sort_values(["sector","season","year"]).reset_index(drop=True)

    print(f"[data_loader] {len(merged)} rows | {merged['sector'].nunique()} sectors (incl. Ngamba proxy) | "
          f"years {merged['year'].min()}-{merged['year'].max()}")
    return rf, mx, mn, merged
